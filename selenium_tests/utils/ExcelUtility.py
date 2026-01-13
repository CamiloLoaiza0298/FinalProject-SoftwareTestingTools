from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import openpyxl

test_results = []

def add_result(
    scenario,
    test_id,
    description,
    steps,
    expected,
    actual,
    status,
    testdata="Not required"
):
    test_results.append({
        "Test Case Scenario": scenario,
        "Test Case ID": test_id,
        "Testcase Description": description,
        "Testcase Steps": steps,
        "Expected Result": expected,
        "Actual Result": actual,
        "Testdata": testdata,
        "Status": status
    })

def write_results_to_excel():
    if not test_results:
        print("No test results to write.")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Define styles
    header_font = Font(bold=True)
    header_color = "5BADD9"
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Write headers
    headers = list(test_results[0].keys())
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.fill = openpyxl.styles.PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    
    # Write data
    for row_num, result in enumerate(test_results, 2):
        for col_num, key in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=result[key])
            cell.border = thin_border
            if key == "Status":
                if result[key] == "Pass":
                    cell.fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                else:
                    cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            if key == "Testcase Steps" or key == "Testdata":
                cell.alignment = Alignment(wrap_text=True)
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = ws.cell(row=1, column=col_num).column_letter
        max_length = len(header)
        for row in range(1, len(test_results) + 2):
            cell_value = str(ws.cell(row=row, column=col_num).value or "")
            max_length = max(max_length, len(cell_value))
        if header in ["Testcase Steps", "Testdata"]:
            ws.column_dimensions[column_letter].width = 65
        else:
            ws.column_dimensions[column_letter].width = max_length + 2
    
    wb.save("test_results.xlsx")
    print("\nTest results written to test_results.xlsx with formatting.")
